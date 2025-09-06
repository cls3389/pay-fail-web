#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理核心服务 - 完整版本
与原始工具保持完全一致的处理逻辑
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import warnings
from datetime import datetime
import traceback
from typing import Dict, List, Tuple, Optional
from config import Config

# 忽略警告
warnings.filterwarnings('ignore')

class ExcelProcessorService:
    """Excel处理服务类 - 与原始版本保持一致"""
    
    def __init__(self):
        self.config = Config()
    
    def process_excel_for_preview(self, input_path: str) -> Dict:
        """
        处理Excel文件并返回预览数据
        """
        try:
            result = {
                'success': False,
                'message': '',
                'preview_data': {},
                'stats': {},
                'errors': []
            }
            
            print(f"📁 正在处理文件用于预览: {input_path}")
            
            # 第1步：读取和验证Excel文件
            df, validation_result = self._read_and_validate_excel(input_path)
            if not validation_result['success']:
                result['errors'] = validation_result['errors']
                result['message'] = validation_result['message']
                return result
            
            result['stats']['原始数据行数'] = len(df)
            result['stats']['检测到的列'] = list(df.columns)
            
            # 第2步：数据预处理
            df = self._preprocess_data(df)
            
            # 第3步：创建透视表
            pivot_table = self._create_pivot_table_full_logic(df)
            
            # 第4步：按直营中心分组预览数据
            preview_data = self._generate_preview_data(pivot_table)
            
            result.update({
                'success': True,
                'message': '数据处理完成',
                'preview_data': preview_data,
                'stats': {
                    **result['stats'],
                    '透视表行数': len(pivot_table),
                    '直营中心数量': len(preview_data),
                    '总金额': float(df['应还款金额'].sum()),
                    '处理时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            })
            
            return result
            
        except Exception as e:
            return {
                'success': False,
                'message': f'处理文件时发生错误: {str(e)}',
                'preview_data': {},
                'stats': {},
                'errors': [str(e), traceback.format_exc()]
            }
    
    def _generate_preview_data(self, pivot_table: pd.DataFrame) -> Dict:
        """生成按直营中心分组的预览数据，包含Excel样式信息"""
        preview_data = {}
        
        if '所属直营中心' not in pivot_table.columns:
            preview_data['default'] = {
                'name': '数据预览',
                'data': pivot_table.drop('所属直营中心', axis=1, errors='ignore').to_dict('records'),
                'columns': list(pivot_table.drop('所属直营中心', axis=1, errors='ignore').columns),
                'total_amount': float(pivot_table['应还款金额'].sum()) if '应还款金额' in pivot_table.columns else 0
            }
            return preview_data
        
        # 按直营中心分组，生成与Excel相同的表格结构
        for center in pivot_table['所属直营中心'].unique():
            center_data = pivot_table[pivot_table['所属直营中心'] == center].copy()
            display_data = center_data.drop('所属直营中心', axis=1)
            
            # 生成Excel风格的表格数据
            table_data = self._generate_excel_style_table(center, display_data)
            
            preview_data[center] = {
                'name': center,
                'excel_table': table_data,  # Excel风格的表格数据
                'raw_data': display_data.to_dict('records'),  # 原始数据（备用）
                'columns': list(display_data.columns),
                'total_amount': float(center_data['应还款金额'].sum()) if '应还款金额' in center_data.columns else 0,
                'row_count': len(center_data)
            }
        
        print(f"   ✅ 生成Excel风格预览数据完成，包含 {len(preview_data)} 个直营中心")
        
        return preview_data
    
    def _generate_excel_style_table(self, center_name: str, data: pd.DataFrame) -> Dict:
        """生成与Excel完全一致的表格结构"""
        table_structure = {
            'center_title': center_name,
            'headers': list(data.columns),
            'rows': [],
            'merge_info': {}  # 合并单元格信息
        }
        
        # 合并规则：只有前两列（团队、业务经理）可以合并，客户姓名和金额不合并
        merge_columns = []
        for i, col in enumerate(data.columns):
            if col in ['所属团队', '所属业务经理']:
                merge_columns.append((i, col))
        
        # 处理数据行
        for idx, row in data.iterrows():
            row_data = []
            for col_idx, col in enumerate(data.columns):
                value = row[col]
                
                # 格式化金额列
                if col == '应还款金额' and isinstance(value, (int, float)):
                    formatted_value = f"{value:,.2f}"
                    # 检查是否需要条件格式（红色高亮）
                    is_highlight = value >= 10000
                    row_data.append({
                        'value': formatted_value,
                        'raw_value': value,
                        'is_amount': True,
                        'is_highlight': is_highlight
                    })
                else:
                    row_data.append({
                        'value': str(value) if value is not None else '',
                        'raw_value': value,
                        'is_amount': False,
                        'is_highlight': False
                    })
            
            table_structure['rows'].append(row_data)
        
        # 计算合并信息（只合并团队和业务经理列）
        merge_groups = {}
        for col_idx, col_name in merge_columns:
            merge_groups[col_name] = []
            current_value = None
            start_row = 0
            
            for row_idx, row_data in enumerate(table_structure['rows']):
                cell_value = row_data[col_idx]['raw_value']
                
                if current_value is None or current_value != cell_value:
                    # 结束上一个合并组
                    if current_value is not None:
                        merge_groups[col_name].append({
                            'value': current_value,
                            'start': start_row,
                            'end': row_idx - 1
                        })
                    
                    # 开始新的合并组
                    current_value = cell_value
                    start_row = row_idx
            
            # 处理最后一个合并组
            if current_value is not None:
                merge_groups[col_name].append({
                    'value': current_value,
                    'start': start_row,
                    'end': len(table_structure['rows']) - 1
                })
        
        table_structure['merge_info'] = merge_groups
        
        # 调试输出合并信息
        for col_name, groups in merge_groups.items():
            print(f"   {col_name} 合并组:")
            for group in groups:
                print(f"     行 {group['start']+1}-{group['end']+1}: {group['value']}")
        
        print(f"   生成表格结构: {len(table_structure['rows'])} 行, 合并列: {list(merge_groups.keys())}")
        
        return table_structure

    def process_excel_file(self, input_path: str, output_dir: str = None) -> Dict:
        """
        处理Excel文件的主要方法 - 完全复制原始逻辑
        """
        try:
            result = {
                'success': False,
                'message': '',
                'output_file': None,
                'stats': {},
                'errors': []
            }
            
            if output_dir is None:
                output_dir = self.config.OUTPUT_FOLDER
            
            print(f"📁 正在处理文件: {input_path}")
            
            # 第1步：读取和验证Excel文件
            df, validation_result = self._read_and_validate_excel(input_path)
            if not validation_result['success']:
                result['errors'] = validation_result['errors']
                result['message'] = validation_result['message']
                return result
            
            result['stats']['原始数据行数'] = len(df)
            result['stats']['检测到的列'] = list(df.columns)
            
            # 第2步：数据预处理
            df = self._preprocess_data(df)
            
            # 第3步：创建透视表（使用原始完整逻辑）
            pivot_table = self._create_pivot_table_full_logic(df)
            
            # 第4步：生成输出文件
            output_file = self._save_to_excel_full_style(df, pivot_table, output_dir)
            
            result.update({
                'success': True,
                'message': '文件处理完成',
                'output_file': output_file,
                'stats': {
                    **result['stats'],
                    '透视表行数': len(pivot_table),
                    '直营中心数量': df['所属直营中心'].nunique() if '所属直营中心' in df.columns else 0,
                    '总金额': float(df['应还款金额'].sum())
                }
            })
            
            return result
            
        except Exception as e:
            return {
                'success': False,
                'message': f'处理文件时发生错误: {str(e)}',
                'output_file': None,
                'stats': {},
                'errors': [str(e), traceback.format_exc()]
            }
    
    def _read_and_validate_excel(self, file_path: str) -> Tuple[pd.DataFrame, Dict]:
        """读取并验证Excel文件"""
        result = {'success': False, 'message': '', 'errors': []}
        
        try:
            if not os.path.exists(file_path):
                result['message'] = f'文件不存在: {file_path}'
                result['errors'].append(result['message'])
                return None, result
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in self.config.ALLOWED_EXTENSIONS:
                result['message'] = f'不支持的文件格式: {file_ext}，请使用 .xlsx 或 .xls 格式'
                result['errors'].append(result['message'])
                return None, result
            
            # 读取Excel文件
            print("   正在读取Excel文件...")
            df = pd.read_excel(file_path)
            print(f"   ✅ 成功读取 {len(df)} 行数据")
            
            if len(df) == 0:
                result['message'] = 'Excel文件为空'
                result['errors'].append(result['message'])
                return None, result
            
            # 显示所有列名
            print(f"   文件包含的列: {list(df.columns)}")
            
            # 验证必要列
            missing_columns = [col for col in self.config.REQUIRED_COLUMNS if col not in df.columns]
            if missing_columns:
                result['message'] = f'缺少必要的列: {missing_columns}'
                result['errors'].append(result['message'])
                return None, result
            
            # 检查是否有客户UID列
            if '客户UID' in df.columns:
                print(f"   ✅ 检测到客户UID列，将用于去重计数")
            else:
                print(f"   ⚠️  未检测到客户UID列，将使用客户姓名去重计数")
            
            result['success'] = True
            result['message'] = f'成功读取 {len(df)} 行数据'
            
            return df, result
            
        except Exception as e:
            result['message'] = f'读取文件失败: {str(e)}'
            result['errors'].append(str(e))
            return None, result
    
    def _preprocess_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据预处理 - 与原始版本保持一致"""
        df = df.copy()
        
        # 处理应还款金额格式
        print("   正在格式化应还款金额...")
        df['应还款金额'] = pd.to_numeric(df['应还款金额'], errors='coerce').fillna(0)
        print(f"   ✅ 应还款金额格式化完成")
        
        # 处理贷后BP逻辑
        if '贷后BP' in df.columns and '所属业务经理' in df.columns:
            print(f"   正在处理贷后BP逻辑...")
            def 获取业务经理(row):
                if pd.notna(row['贷后BP']) and str(row['贷后BP']).strip() not in ['无', '']:
                    return row['贷后BP']
                else:
                    return row['所属业务经理']
            df['所属业务经理'] = df.apply(获取业务经理, axis=1)
            print(f"   ✅ 贷后BP逻辑处理完成")
        
        # 应用排序
        df = self._sort_data(df)
        
        return df
    
    def _sort_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据排序 - 完整的原始逻辑"""
        print("   正在对原始数据进行排序...")
        try:
            from pypinyin import pinyin, Style
            df['拼音排序键'] = df['所属直营中心'].apply(
                lambda x: ''.join([p[0] for p in pinyin(str(x), style=Style.NORMAL)])
            )
            df = df.sort_values('拼音排序键', ascending=True)
            df = df.drop('拼音排序键', axis=1)
            print("   ✅ 使用拼音排序完成")
            print("   排序后的直营中心顺序：")
            for i, 直营中心 in enumerate(df['所属直营中心'].unique()[:5]):
                print(f"     {i+1}. {直营中心}")
        except ImportError:
            df = df.sort_values('所属直营中心', ascending=True)
            print("   ⚠️  使用Unicode排序（建议安装pypinyin获得更好的中文排序）")
        
        return df
    
    def _create_pivot_table_full_logic(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建数据透视表 - 完整的原始逻辑"""
        print("   正在创建数据透视表...")
        
        # 透视表行字段
        透视表行字段 = ['所属团队', '所属业务经理', '客户姓名']
        存在的透视表行字段 = [字段 for 字段 in 透视表行字段 if 字段 in df.columns]
        
        # 保留直营中心信息用于分组显示
        if '所属直营中心' in df.columns:
            存在的透视表行字段.insert(0, '所属直营中心')
        
        # 创建透视表
        透视表 = pd.pivot_table(
            df,
            values=['应还款金额'],
            index=存在的透视表行字段,
            aggfunc={'应还款金额': 'sum'},
            fill_value=0
        ).reset_index()
        
        # 扁平化列名
        if isinstance(透视表.columns, pd.MultiIndex):
            透视表.columns = [col[1] if col[1] else col[0] for col in 透视表.columns.values]
        
        print(f"   ✅ 基础透视表创建完成，共 {len(透视表)} 行")
        
        # 应用去重计数排序逻辑
        print("   正在应用去重计数排序逻辑...")
        
        if '所属直营中心' in df.columns:
            # 计算团队和业务经理的去重客户数量
            print("     正在计算团队和业务经理的去重客户数量...")
            
            # 使用客户UID或客户姓名进行去重
            去重字段 = '客户UID' if '客户UID' in df.columns else '客户姓名'
            
            # 按团队分组，计算去重后的客户数量
            团队统计 = df.groupby('所属团队')[去重字段].nunique().reset_index()
            团队统计.columns = ['所属团队', '团队客户数量']
            
            # 按业务经理分组，计算去重后的客户数量
            业务经理统计 = df.groupby('所属业务经理')[去重字段].nunique().reset_index()
            业务经理统计.columns = ['所属业务经理', '业务经理客户数量']
            
            # 将统计信息合并到透视表
            透视表 = 透视表.merge(团队统计, on='所属团队', how='left')
            透视表 = 透视表.merge(业务经理统计, on='所属业务经理', how='left')
            
            # 创建团队排序键（贷后BP团队置底）
            def 获取团队排序键(团队名):
                return 999999 if '贷后BP团队' in str(团队名) else 0
            
            透视表['团队排序键'] = 透视表['所属团队'].apply(获取团队排序键)
            透视表['团队客户数量'] = 透视表['团队客户数量'].astype(int)
            透视表['业务经理客户数量'] = 透视表['业务经理客户数量'].astype(int)
            
            # 创建团队到直营中心的映射和排序
            团队直营中心映射 = df[['所属团队', '所属直营中心']].drop_duplicates().set_index('所属团队')['所属直营中心'].to_dict()
            透视表['所属直营中心'] = 透视表['所属团队'].map(团队直营中心映射)
            
            # 创建排序键，确保直营中心按原始顺序
            直营中心顺序 = df['所属直营中心'].unique()
            直营中心顺序字典 = {直营中心: i for i, 直营中心 in enumerate(直营中心顺序)}
            透视表['直营中心顺序键'] = 透视表['所属直营中心'].map(直营中心顺序字典)
            
            # 完整排序逻辑
            透视表 = 透视表.sort_values([
                '直营中心顺序键',        # 第一优先级：保持直营中心原始顺序
                '团队客户数量',          # 第二优先级：团队去重客户数量（降序）
                '团队排序键',            # 第三优先级：团队排序键（贷后BP团队置底）
                '所属团队',              # 第四优先级：团队名称（降序，确保相同团队聚集）
                '业务经理客户数量',      # 第五优先级：业务经理去重客户数量（降序）
                '所属业务经理',          # 第六优先级：业务经理名称
                '应还款金额'             # 第七优先级：应还款金额（降序）
            ], ascending=[True, False, True, False, False, True, False])
            
            # 删除临时排序键，但保留直营中心信息用于后续处理
            透视表 = 透视表.drop(['直营中心顺序键', '团队客户数量', '业务经理客户数量', '团队排序键'], axis=1)
        
        print(f"   ✅ 透视表排序逻辑应用完成")
        
        return 透视表
    
    def _save_to_excel_full_style(self, df: pd.DataFrame, pivot_table: pd.DataFrame, output_dir: str) -> str:
        """保存到Excel文件 - 完整样式"""
        # 生成输出文件名
        current_time = datetime.now()
        sheet_name = f"{current_time.month:02d}{current_time.day:02d}{current_time.hour:02d}{current_time.minute:02d}"
        
        # 生成输出文件名（不删除原文件）
        base_name = "扣款失败信息处理"
        timestamp = current_time.strftime('%Y%m%d_%H%M%S')
        output_filename = f"{base_name}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        # 保存Excel文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 透视表工作表
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 原始数据工作表  
            df.to_excel(writer, sheet_name='原始数据', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            pivot_ws = workbook[sheet_name]
            raw_ws = workbook['原始数据']
            
            # 应用完整样式
            self._apply_pivot_table_style_full(pivot_ws, pivot_table)
            self._apply_raw_data_style_full(raw_ws, df)
        
        print(f"   ✅ 文件保存完成: {output_path}")
        
        return output_path
    
    def _apply_pivot_table_style_full(self, ws, pivot_table: pd.DataFrame):
        """应用完整的透视表样式"""
        if '所属直营中心' in pivot_table.columns:
            self._apply_pivot_style_with_center_title_full(ws, pivot_table)
        else:
            self._apply_basic_pivot_style_full(ws, pivot_table)
    
    def _apply_pivot_style_with_center_title_full(self, ws, pivot_table: pd.DataFrame):
        """应用带直营中心标题的完整透视表样式"""
        print("   正在应用透视表样式（直营中心标题模式）...")
        
        # 移除直营中心列创建显示用透视表
        显示透视表 = pivot_table.drop('所属直营中心', axis=1).copy()
        表头 = list(显示透视表.columns)
        
        # 清空工作表
        ws.delete_rows(1, ws.max_row)
        
        # 按直营中心分组插入数据
        当前行 = 1
        当前直营中心 = None
        
        for idx, row in pivot_table.iterrows():
            直营中心 = row['所属直营中心']
            
            # 新直营中心，插入标题
            if 直营中心 != 当前直营中心:
                if 当前直营中心 is not None:
                    当前行 += 1  # 空行分隔
                
                # 插入直营中心标题
                ws.cell(row=当前行, column=1, value=直营中心)
                ws.merge_cells(start_row=当前行, end_row=当前行, 
                             start_column=1, end_column=len(表头))
                当前行 += 1
                
                # 插入表头
                for col_idx, header in enumerate(表头, 1):
                    ws.cell(row=当前行, column=col_idx, value=header)
                当前行 += 1
                
                当前直营中心 = 直营中心
            
            # 插入数据行
            for col_idx, header in enumerate(表头, 1):
                if header in row:
                    ws.cell(row=当前行, column=col_idx, value=row[header])
            当前行 += 1
        
        # 应用基础样式和合并
        self._apply_excel_styles_full(ws, len(表头))
        self._apply_cell_merge_full(ws, len(表头))
        
        print("   ✅ 透视表样式应用完成（直营中心标题模式）")
    
    def _apply_basic_pivot_style_full(self, ws, pivot_table: pd.DataFrame):
        """应用基础透视表样式"""
        self._apply_excel_styles_full(ws, len(pivot_table.columns))
    
    def _apply_raw_data_style_full(self, ws, df: pd.DataFrame):
        """应用原始数据完整样式"""
        self._apply_excel_styles_full(ws, len(df.columns))
        
        # 添加筛选功能
        ws.auto_filter.ref = ws.dimensions
        
        print("   ✅ 原始数据样式应用完成")
    
    def _apply_excel_styles_full(self, ws, num_columns: int):
        """应用完整的Excel样式"""
        print("     正在应用基础样式...")
        
        # 定义样式
        标题字体 = Font(name='微软雅黑', size=11, bold=True, color='000000')
        内容字体 = Font(name='微软雅黑', size=10, color='000000')
        直营中心标题字体 = Font(name='微软雅黑', size=14, bold=True, color='000000')
        居中对齐 = Alignment(horizontal='center', vertical='center')
        边框样式 = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 应用样式到所有单元格
        for row in range(1, ws.max_row + 1):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                
                # 检查是否为直营中心标题行（通过合并单元格判断）
                if col == 1 and cell.value and cell.coordinate in ws.merged_cells:
                    # 直营中心标题样式
                    cell.font = 直营中心标题字体
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                    ws.row_dimensions[row].height = 25  # 与CSS前端保持一致
                elif any(keyword in str(cell.value or '') for keyword in ['所属团队', '所属业务经理', '客户姓名', '应还款金额']):
                    # 表头样式
                    cell.font = 标题字体
                    cell.alignment = 居中对齐
                    cell.border = 边框样式
                    ws.row_dimensions[row].height = 22
                elif cell.value is None or str(cell.value).strip() == '':
                    # 空行
                    ws.row_dimensions[row].height = 15
                else:
                    # 数据行样式
                    cell.font = 内容字体
                    cell.alignment = 居中对齐
                    cell.border = 边框样式
                    ws.row_dimensions[row].height = 20
                
                # 金额格式化
                if col == num_columns:  # 假设最后一列是金额列
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.number_format = '#,##0.00'
                        # 条件格式
                        if cell.value >= self.config.FORMAT_CONFIG['金额阈值']:
                            cell.fill = PatternFill(
                                start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'
                            )
                            cell.font = Font(name='微软雅黑', size=10, color='8B0000')
        
        # 智能调整列宽
        print("     正在智能调整列宽...")
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            if col == 1:
                ws.column_dimensions[column_letter].width = 18
            elif col == 2:
                ws.column_dimensions[column_letter].width = 16
            elif col == 3:
                ws.column_dimensions[column_letter].width = 14
            elif col == num_columns:
                ws.column_dimensions[column_letter].width = 15
            else:
                ws.column_dimensions[column_letter].width = 12
        
        print("     ✅ 基础样式应用完成")
    
    def _apply_cell_merge_full(self, ws, 列数: int):
        """应用完整的单元格合并功能"""
        print("     正在应用单元格合并...")
        
        # 定义样式
        内容字体 = Font(name='微软雅黑', size=10, color='000000')
        居中对齐 = Alignment(horizontal='center', vertical='center')
        边框样式 = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 对前几列进行合并（不合并最后一列金额列）
        for col in range(1, 列数):
            current_value = None
            start_row = 1
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value
                
                # 跳过标题行
                是否标题行 = (col == 1 and cell_value and ws.cell(row=row, column=2).value is None and 
                          (列数 < 3 or ws.cell(row=row, column=3).value is None))
                是否表头行 = (cell_value and any(keyword in str(cell_value) for keyword in 
                          ['所属团队', '所属业务经理', '客户姓名', '应还款金额']))
                
                if 是否标题行 or 是否表头行:
                    if current_value is not None and start_row < row - 1:
                        ws.merge_cells(start_row=start_row, end_row=row-1, start_column=col, end_column=col)
                        merged_cell = ws.cell(row=start_row, column=col)
                        merged_cell.font = 内容字体
                        merged_cell.alignment = 居中对齐
                        merged_cell.border = 边框样式
                    
                    current_value = None
                    start_row = row + 1
                    continue
                
                if cell_value != current_value:
                    if current_value is not None and start_row < row - 1:
                        ws.merge_cells(start_row=start_row, end_row=row-1, start_column=col, end_column=col)
                        merged_cell = ws.cell(row=start_row, column=col)
                        merged_cell.font = 内容字体
                        merged_cell.alignment = 居中对齐
                        merged_cell.border = 边框样式
                    
                    current_value = cell_value
                    start_row = row
            
            # 处理最后一组
            if current_value is not None and start_row < ws.max_row:
                ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=col, end_column=col)
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.font = 内容字体
                merged_cell.alignment = 居中对齐
                merged_cell.border = 边框样式
        
        print("     ✅ 单元格合并完成")

# 创建全局服务实例
excel_service = ExcelProcessorService()