#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理独立版本 - 完全不依赖Web框架
专用于本地exe打包，输出行为与原始版本一致
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import warnings
from datetime import datetime
import traceback
from typing import Dict, List, Tuple, Optional
from pathlib import Path

# 忽略警告
warnings.filterwarnings('ignore')

class StandaloneExcelProcessor:
    """独立Excel处理器 - 与原始版本行为完全一致"""
    
    def __init__(self):
        # 硬编码配置，与原始版本一致
        self.required_columns = ['应还款金额', '所属直营中心', '所属团队', '所属业务经理', '客户姓名']
        self.optional_columns = ['客户UID', '贷后BP']
        
        # 排序配置
        self.sort_config = {
            '团队排序': True,
            '业务经理排序': True,
            '金额排序': True,
            '贷后BP团队置底': True
        }
        
        # 条件格式配置
        self.format_config = {
            '金额阈值': 10000,
            '浅红填充色': 'FFE6E6',
            '深红色文本': 'CC0000'
        }
    
    def process_excel_file(self, input_path: str) -> Dict:
        """
        处理Excel文件的主要方法 - 输出到当前目录
        """
        try:
            result = {
                'success': False,
                'message': '',
                'output_file': None,
                'stats': {},
                'errors': []
            }
            
            print(f"📁 正在处理文件: {input_path}")
            
            # 第1步：读取和验证Excel文件
            df, validation_result = self._read_and_validate_excel(input_path)
            if not validation_result['success']:
                result['errors'] = validation_result['errors']
                result['message'] = validation_result['message']
                return result
            
            result['stats']['原始数据行数'] = len(df)
            result['stats']['检测到的列'] = list(df.columns)
            
            # 第2步：数据预处理
            df = self._preprocess_data(df)
            
            # 第3步：创建透视表
            pivot_table = self._create_pivot_table(df)
            
            # 第4步：生成输出文件 - 输出到当前目录
            output_file = self._save_to_excel_current_dir(df, pivot_table)
            
            result.update({
                'success': True,
                'message': '文件处理完成',
                'output_file': output_file,
                'stats': {
                    **result['stats'],
                    '透视表行数': len(pivot_table),
                    '直营中心数量': df['所属直营中心'].nunique() if '所属直营中心' in df.columns else 0,
                    '总金额': float(df['应还款金额'].sum())
                }
            })
            
            return result
            
        except Exception as e:
            print(f"❌ 处理过程中发生错误: {e}")
            result.update({
                'success': False,
                'message': str(e),
                'errors': [str(e), traceback.format_exc()]
            })
            return result
    
    def _read_and_validate_excel(self, input_path: str) -> Tuple[pd.DataFrame, Dict]:
        """读取和验证Excel文件"""
        try:
            print("   正在读取Excel文件...")
            df = pd.read_excel(input_path)
            print(f"   ✅ 成功读取 {len(df)} 行数据")
            
            # 显示所有列名
            print(f"   文件包含的列: {list(df.columns)}")
            
            # 检查必要的列
            missing_columns = [col for col in self.required_columns if col not in df.columns]
            if missing_columns:
                error_msg = f"缺少必要的列: {missing_columns}"
                print(f"   ❌ {error_msg}")
                return df, {
                    'success': False,
                    'message': error_msg,
                    'errors': [f"必需列 {col} 不存在" for col in missing_columns]
                }
            
            # 检查是否有客户UID列
            if '客户UID' in df.columns:
                print(f"   ✅ 检测到客户UID列，将用于去重计数")
            else:
                print(f"   ⚠️  未检测到客户UID列，将使用客户姓名去重计数")
            
            return df, {'success': True, 'message': 'Excel文件读取成功'}
            
        except Exception as e:
            error_msg = f"读取Excel文件失败: {e}"
            print(f"   ❌ {error_msg}")
            return pd.DataFrame(), {
                'success': False,
                'message': error_msg,
                'errors': [str(e)]
            }
    
    def _preprocess_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据预处理"""
        print("   正在进行数据预处理...")
        
        # 处理应还款金额格式
        print("   正在格式化应还款金额...")
        df['应还款金额'] = pd.to_numeric(df['应还款金额'], errors='coerce').fillna(0)
        print(f"   ✅ 应还款金额格式化完成")
        
        # 处理贷后BP逻辑
        if '贷后BP' in df.columns and '所属业务经理' in df.columns:
            print(f"   正在处理贷后BP逻辑...")
            df = df.copy()  # 创建副本避免警告
            def 获取业务经理(row):
                if pd.notna(row['贷后BP']) and str(row['贷后BP']).strip() not in ['无', '']:
                    return row['贷后BP']
                else:
                    return row['所属业务经理']
            
            df['所属业务经理'] = df.apply(获取业务经理, axis=1)
            print(f"   ✅ 贷后BP逻辑处理完成")
        
        print("   ✅ 数据预处理完成")
        return df
    
    def _create_pivot_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """创建数据透视表"""
        print("   正在创建数据透视表...")
        
        # 透视表行字段
        pivot_index = ['所属团队', '所属业务经理', '客户姓名']
        available_index = [field for field in pivot_index if field in df.columns]
        
        if not available_index:
            raise ValueError("无法创建透视表：缺少必要的分组字段")
        
        # 创建透视表
        pivot_table = df.groupby(available_index).agg({
            '应还款金额': 'sum'
        }).reset_index()
        
        # 智能排序逻辑
        if len(pivot_table) > 1:
            # 去重字段（优先使用客户UID）
            unique_field = '客户UID' if '客户UID' in df.columns else '客户姓名'
            
            # 按团队分组，计算去重后的客户数量
            team_stats = df.groupby('所属团队')[unique_field].nunique().reset_index()
            team_stats.columns = ['所属团队', '团队客户数量']
            
            # 按业务经理分组，计算去重后的客户数量  
            manager_stats = df.groupby('所属业务经理')[unique_field].nunique().reset_index()
            manager_stats.columns = ['所属业务经理', '业务经理客户数量']
            
            # 将统计信息合并到透视表
            pivot_table = pivot_table.merge(team_stats, on='所属团队', how='left')
            pivot_table = pivot_table.merge(manager_stats, on='所属业务经理', how='left')
            
            # 创建团队排序键（贷后BP团队置底）
            def 获取团队排序键(团队名):
                return 999999 if '贷后BP团队' in str(团队名) else 0
            
            pivot_table['团队排序键'] = pivot_table['所属团队'].apply(获取团队排序键)
            pivot_table['团队客户数量'] = pivot_table['团队客户数量'].astype(int)
            pivot_table['业务经理客户数量'] = pivot_table['业务经理客户数量'].astype(int)
            
            # 创建团队到直营中心的映射
            team_center_mapping = df[['所属团队', '所属直营中心']].drop_duplicates().set_index('所属团队')['所属直营中心'].to_dict()
            pivot_table['所属直营中心'] = pivot_table['所属团队'].map(team_center_mapping)
            
            # 创建排序键，确保直营中心按原始顺序
            center_order = df['所属直营中心'].unique()
            center_order_dict = {center: i for i, center in enumerate(center_order)}
            pivot_table['直营中心顺序键'] = pivot_table['所属直营中心'].map(center_order_dict)
            
            # 完整排序逻辑
            pivot_table = pivot_table.sort_values([
                '直营中心顺序键',        # 第一优先级：保持直营中心原始顺序
                '团队客户数量',          # 第二优先级：团队去重客户数量（降序）
                '团队排序键',            # 第三优先级：团队排序键（贷后BP团队置底）
                '所属团队',              # 第四优先级：团队名称
                '业务经理客户数量',      # 第五优先级：业务经理去重客户数量（降序）
                '所属业务经理',          # 第六优先级：业务经理名称
                '应还款金额'             # 第七优先级：应还款金额（降序）
            ], ascending=[True, False, True, False, False, True, False])
            
            # 删除临时排序键，但保留直营中心信息用于后续处理
            pivot_table = pivot_table.drop(['直营中心顺序键', '团队客户数量', '业务经理客户数量', '团队排序键'], axis=1)
        
        print(f"   ✅ 透视表创建完成，共{len(pivot_table)}行")
        return pivot_table
    
    def _save_to_excel_current_dir(self, df: pd.DataFrame, pivot_table: pd.DataFrame) -> str:
        """保存到Excel文件 - 输出到当前目录，与原始版本一致"""
        # 生成输出文件名（与原始版本一致）
        current_time = datetime.now()
        sheet_name = f"{current_time.month:02d}{current_time.day:02d}{current_time.hour:02d}{current_time.minute:02d}"
        
        # 输出到当前目录，文件名格式与原始版本一致
        base_name = "扣款失败信息处理结果"
        timestamp = current_time.strftime('%Y%m%d_%H%M%S')
        output_filename = f"{base_name}_{timestamp}.xlsx"
        output_path = os.path.abspath(output_filename)  # 当前目录的绝对路径
        
        print(f"   正在保存文件到: {output_path}")
        
        # 保存Excel文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 透视表工作表
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 原始数据工作表  
            df.to_excel(writer, sheet_name='原始数据', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            pivot_ws = workbook[sheet_name]
            raw_ws = workbook['原始数据']
            
            # 应用样式（简化版本，专注核心功能）
            self._apply_basic_styles(pivot_ws, pivot_table)
            self._apply_basic_styles(raw_ws, df)
        
        print(f"   ✅ 文件保存完成: {output_filename}")
        
        return output_path
    
    def _apply_basic_styles(self, ws, data_df: pd.DataFrame):
        """应用基础样式"""
        # 设置字体和对齐
        header_font = Font(name='微软雅黑', bold=True, size=11)
        content_font = Font(name='微软雅黑', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for col in range(1, len(data_df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            # 表头背景色
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # 应用数据样式
        for row in range(2, len(data_df) + 2):
            for col in range(1, len(data_df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = content_font
                cell.border = border
                
                # 金额列居右对齐并添加千分位分隔符
                if '应还款金额' in data_df.columns and col == data_df.columns.get_loc('应还款金额') + 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.00'
                    
                    # 条件格式：大于阈值的金额标红
                    if isinstance(cell.value, (int, float)) and cell.value > self.format_config['金额阈值']:
                        cell.fill = PatternFill(start_color=self.format_config['浅红填充色'], 
                                              end_color=self.format_config['浅红填充色'], 
                                              fill_type='solid')
                        cell.font = Font(name='微软雅黑', size=10, color=self.format_config['深红色文本'])
                else:
                    cell.alignment = center_alignment
        
        # 自动调整列宽
        for col in range(1, len(data_df.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(data_df) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            # 设置列宽，中文字符按2个字符宽度计算
            adjusted_width = min(max_length * 1.2 + 2, 50)  # 最大宽度限制
            ws.column_dimensions[column_letter].width = adjusted_width


# 创建全局实例
standalone_processor = StandaloneExcelProcessor()

def process_excel_file(input_path: str) -> Dict:
    """处理Excel文件的全局函数"""
    return standalone_processor.process_excel_file(input_path)
